"""
Test Case Generator for LeetCode Problems

Parses examples from problem descriptions and generates additional edge cases
based on constraint analysis. Outputs tests.json in each problem directory.

Test format (language-agnostic, mirrors LeetCode judge I/O):
{
  "problem_id": 1,
  "title": "Two Sum",
  "test_cases": [
    {
      "input": {"nums": [2,7,11,15], "target": 9},
      "expected": [0,1],
      "source": "example"
    },
    ...
  ]
}
"""

import json
import os
import re
import sys
import io
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Example parsing
# ---------------------------------------------------------------------------

def extract_examples(description):
    """Extract Input/Output pairs from a problem description."""
    if not description:
        return []

    examples = []
    # Split into example blocks
    # Pattern: "Example N:" followed by content until next "Example" or "Constraints" or end
    example_blocks = re.split(r'Example\s*\d*\s*:', description)

    for block in example_blocks[1:]:  # skip text before first example
        input_match = re.search(
            r'Input\s*:\s*(.+?)(?=Output\s*:)',
            block,
            re.DOTALL
        )
        output_match = re.search(
            r'Output\s*:\s*(.+?)(?=Explanation|Example|\n\n|\nNote|\nConstraint|$)',
            block,
            re.DOTALL
        )

        if input_match and output_match:
            raw_input = input_match.group(1).strip()
            raw_output = output_match.group(1).strip()

            parsed_input = parse_input_params(raw_input)
            parsed_output = parse_value(raw_output)

            if parsed_input is not None and parsed_output is not None:
                examples.append({
                    "input": parsed_input,
                    "expected": parsed_output,
                    "source": "example",
                })

    return examples


def parse_input_params(raw_input):
    """Parse 'param1 = value1, param2 = value2' into a dict.

    Handles multi-line values (matrices) and nested structures.
    """
    # Clean up whitespace
    raw_input = raw_input.replace('\r', '')

    params = {}
    # Strategy: find 'name = ' patterns, then capture values between them
    # We need to handle cases like: nums = [1,2,3], target = 9
    # And: board = [["5","3",...],["6",...]]

    # Find all parameter names and their start positions
    param_pattern = re.finditer(r'(\w+)\s*=\s*', raw_input)
    param_starts = []
    for m in param_pattern:
        param_starts.append((m.group(1), m.end()))

    if not param_starts:
        # No named params - might be a single unnamed input
        val = parse_value(raw_input.strip())
        if val is not None:
            return {"arg0": val}
        return None

    for i, (name, start) in enumerate(param_starts):
        # Value extends to the start of next param declaration, or end of string
        if i + 1 < len(param_starts):
            # Find the comma/separator before the next param name
            next_name, next_start = param_starts[i + 1]
            # Look backwards from next param's '=' to find where value ends
            # The pattern is: value, next_name = ...
            # Find the last comma before next_name that's outside brackets
            end = raw_input.rfind(next_name, start, next_start)
            # Go back to find the comma separator
            val_str = raw_input[start:end].strip()
            # Remove trailing comma
            val_str = val_str.rstrip(',').strip()
        else:
            val_str = raw_input[start:].strip()

        parsed = parse_value(val_str)
        if parsed is not None:
            params[name] = parsed
        else:
            params[name] = val_str  # Keep raw if unparseable

    return params if params else None


def parse_value(val_str):
    """Parse a LeetCode-format value into a Python object."""
    val_str = val_str.strip()

    # Remove trailing periods or commas
    val_str = val_str.rstrip('.,;')
    val_str = val_str.strip()

    if not val_str:
        return None

    # Handle special LeetCode values
    if val_str.lower() == 'true':
        return True
    if val_str.lower() == 'false':
        return False
    if val_str.lower() == 'null':
        return None

    # Handle quoted strings
    if (val_str.startswith('"') and val_str.endswith('"')) or \
       (val_str.startswith("'") and val_str.endswith("'")):
        return val_str[1:-1]

    # Try direct JSON parse (handles arrays, numbers, nested arrays)
    try:
        result = json.loads(val_str)
        return result
    except (json.JSONDecodeError, ValueError):
        pass

    # Handle arrays with quotes that aren't valid JSON
    # e.g., ["5","3","."] or [["a","b"],["c","d"]]
    try:
        # Replace single quotes with double quotes for JSON compat
        fixed = val_str.replace("'", '"')
        result = json.loads(fixed)
        return result
    except (json.JSONDecodeError, ValueError):
        pass

    # Try as number
    try:
        if '.' in val_str:
            return float(val_str)
        return int(val_str)
    except ValueError:
        pass

    # Return as string if nothing else works
    # But filter out multi-line junk
    if '\n' in val_str and len(val_str) > 200:
        return None
    return val_str


# ---------------------------------------------------------------------------
# Constraint parsing
# ---------------------------------------------------------------------------

def parse_constraints(description):
    """Extract constraints from description text."""
    if not description:
        return {}

    constraints = {}

    # Find constraints section
    constraint_match = re.search(
        r'Constraints?\s*:(.*?)(?=Follow[ -]?up|$)',
        description,
        re.DOTALL | re.IGNORECASE,
    )
    if not constraint_match:
        return constraints

    text = constraint_match.group(1)

    # Parse common constraint patterns
    # Array length: 1 <= nums.length <= 10^4
    for m in re.finditer(
        r'(\d+)\s*<=\s*(\w+)(?:\.length|\.size)\s*<=\s*(\d+(?:\^?\d*)?)', text
    ):
        min_val = int(m.group(1))
        param = m.group(2)
        max_val = _parse_num(m.group(3))
        constraints[f"{param}_length_min"] = min_val
        constraints[f"{param}_length_max"] = max_val

    # Value bounds: -10^9 <= nums[i] <= 10^9
    for m in re.finditer(
        r'(-?\d+(?:\^?\d*)?)\s*<=\s*(\w+)(?:\[.?\])?\s*<=\s*(-?\d+(?:\^?\d*)?)', text
    ):
        min_val = _parse_num(m.group(1))
        param = m.group(2)
        max_val = _parse_num(m.group(3))
        constraints[f"{param}_val_min"] = min_val
        constraints[f"{param}_val_max"] = max_val

    # String length: 1 <= s.length <= 10^5
    for m in re.finditer(
        r'(\d+)\s*<=\s*(\w+)\.length\s*<=\s*(\d+(?:\^?\d*)?)', text
    ):
        min_val = int(m.group(1))
        param = m.group(2)
        max_val = _parse_num(m.group(3))
        constraints[f"{param}_length_min"] = min_val
        constraints[f"{param}_length_max"] = max_val

    # n, m, k bounds: 1 <= n <= 10^5
    for m in re.finditer(
        r'(-?\d+(?:\^?\d*)?)\s*<=\s*([nmkNMK])\s*<=\s*(-?\d+(?:\^?\d*)?)', text
    ):
        min_val = _parse_num(m.group(1))
        param = m.group(2)
        max_val = _parse_num(m.group(3))
        constraints[f"{param}_min"] = min_val
        constraints[f"{param}_max"] = max_val

    return constraints


def _parse_num(s):
    """Parse a number that might contain ^ for exponent."""
    s = s.strip()
    if '^' in s:
        parts = s.split('^')
        try:
            return int(parts[0]) ** int(parts[1])
        except (ValueError, IndexError):
            pass
    if '10' in s:
        exp_match = re.match(r'10\^?(\d+)', s.replace(' ', ''))
        if exp_match:
            return 10 ** int(exp_match.group(1))
    try:
        return int(s)
    except ValueError:
        return 0


# ---------------------------------------------------------------------------
# Edge case generation
# ---------------------------------------------------------------------------

def infer_param_types(examples):
    """Infer the type of each parameter from example data."""
    if not examples:
        return {}

    types = {}
    first_input = examples[0].get("input", {})

    for param, value in first_input.items():
        if isinstance(value, list):
            if value and isinstance(value[0], list):
                types[param] = "matrix"
            elif value and isinstance(value[0], str):
                types[param] = "string_array"
            else:
                types[param] = "array"
        elif isinstance(value, str):
            types[param] = "string"
        elif isinstance(value, bool):
            types[param] = "bool"
        elif isinstance(value, float):
            types[param] = "float"
        elif isinstance(value, int):
            types[param] = "int"
        else:
            types[param] = "unknown"

    return types


def generate_edge_cases(examples, constraints, param_types):
    """Generate edge cases based on parameter types and constraints."""
    if not examples or not param_types:
        return []

    edge_cases = []
    first_input = examples[0].get("input", {})
    params = list(first_input.keys())

    # Strategy: for each parameter, try boundary values while keeping others at example defaults
    for param in params:
        ptype = param_types.get(param, "unknown")

        if ptype == "array":
            edge_cases.extend(_array_edge_cases(param, first_input, constraints))
        elif ptype == "string":
            edge_cases.extend(_string_edge_cases(param, first_input, constraints))
        elif ptype == "int":
            edge_cases.extend(_int_edge_cases(param, first_input, constraints))
        elif ptype == "float":
            edge_cases.extend(_float_edge_cases(param, first_input, constraints))

    return edge_cases


def _make_case(base_input, overrides, source="edge_case"):
    """Create a test case with overrides applied to base input."""
    new_input = dict(base_input)
    new_input.update(overrides)
    return {
        "input": new_input,
        "expected": "__VERIFY__",
        "source": source,
    }


def _array_edge_cases(param, base_input, constraints):
    """Generate edge cases for array parameters."""
    cases = []
    min_len = constraints.get(f"{param}_length_min", 0)
    val_min = constraints.get(f"{param}_val_min",
                              constraints.get("nums_val_min", -100))
    val_max = constraints.get(f"{param}_val_max",
                              constraints.get("nums_val_max", 100))

    original = base_input.get(param, [])
    elem_example = original[0] if original else 0

    # Minimum length array
    if min_len <= 1:
        cases.append(_make_case(base_input, {param: [elem_example]}))
    elif min_len == 2:
        cases.append(_make_case(base_input, {param: [elem_example, elem_example]}))

    # Two-element array (if different from min)
    if min_len <= 2:
        cases.append(_make_case(base_input, {param: [val_min, val_max]}))

    # All same elements
    if len(original) >= 2:
        cases.append(_make_case(base_input, {param: [elem_example] * len(original)}))

    # Sorted ascending
    if len(original) >= 3:
        n = len(original)
        cases.append(_make_case(base_input, {param: list(range(1, n + 1))}))

    # Sorted descending
    if len(original) >= 3:
        n = len(original)
        cases.append(_make_case(base_input, {param: list(range(n, 0, -1))}))

    # Boundary values
    if val_min is not None and val_max is not None:
        n = max(min_len, 2)
        cases.append(_make_case(base_input, {param: [val_min] * n}))
        cases.append(_make_case(base_input, {param: [val_max] * n}))

    return cases


def _string_edge_cases(param, base_input, constraints):
    """Generate edge cases for string parameters."""
    cases = []
    min_len = constraints.get(f"{param}_length_min", 0)
    original = base_input.get(param, "")

    # Determine character set from original
    if original:
        char = original[0]
    else:
        char = "a"

    # Single character
    if min_len <= 1:
        cases.append(_make_case(base_input, {param: char}))

    # All same character
    if len(original) >= 2:
        cases.append(_make_case(base_input, {param: char * len(original)}))

    # Two characters alternating
    if min_len <= 2:
        other_char = chr(ord(char) + 1) if char < 'z' else 'a'
        n = max(len(original), 4)
        cases.append(_make_case(base_input, {param: (char + other_char) * (n // 2)}))

    # Palindrome
    if len(original) >= 3:
        half = original[:len(original) // 2]
        cases.append(_make_case(base_input, {param: half + half[::-1]}))

    return cases


def _int_edge_cases(param, base_input, constraints):
    """Generate edge cases for integer parameters."""
    cases = []
    val_min = constraints.get(f"{param}_val_min",
                              constraints.get(f"{param}_min", None))
    val_max = constraints.get(f"{param}_val_max",
                              constraints.get(f"{param}_max", None))

    if val_min is not None:
        cases.append(_make_case(base_input, {param: val_min}))
    if val_max is not None:
        cases.append(_make_case(base_input, {param: val_max}))

    # Zero (if in range)
    if (val_min is not None and val_min <= 0 and
            val_max is not None and val_max >= 0):
        cases.append(_make_case(base_input, {param: 0}))

    # One
    if (val_min is not None and val_min <= 1 and
            val_max is not None and val_max >= 1):
        cases.append(_make_case(base_input, {param: 1}))

    return cases


def _float_edge_cases(param, base_input, constraints):
    """Generate edge cases for float parameters."""
    cases = _int_edge_cases(param, base_input, constraints)
    # Add 0.5 boundary
    val_min = constraints.get(f"{param}_val_min", 0)
    val_max = constraints.get(f"{param}_val_max", 100)
    if val_min is not None and val_max is not None:
        mid = (val_min + val_max) / 2
        cases.append(_make_case(base_input, {param: mid}))
    return cases


def deduplicate_cases(cases):
    """Remove duplicate test cases based on input."""
    seen = set()
    unique = []
    for case in cases:
        key = json.dumps(case["input"], sort_keys=True, default=str)
        if key not in seen:
            seen.add(key)
            unique.append(case)
    return unique


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def get_dir_name(num, title):
    """Generate directory name matching the format used by the directory creator."""
    slug = re.sub(r'[^a-z0-9]+', '_', title.lower()).strip('_')
    return f"{int(num)}_{slug}"


def main():
    print("=" * 60)
    print("LeetCode Test Case Generator")
    print("=" * 60)

    wb = load_workbook("leetcode_problems.xlsx")
    ws = wb.active

    total = 0
    success = 0
    no_examples = 0
    with_edge = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        num, title, difficulty, description, topics, hints, link, premium = row
        if premium == "Yes":
            continue

        total += 1
        dir_name = get_dir_name(num, title)
        dir_path = os.path.join("problems", dir_name)

        if not os.path.exists(dir_path):
            continue

        # Parse examples from description
        examples = extract_examples(description)

        # Parse constraints
        constraints = parse_constraints(description)

        # Infer types and generate edge cases
        edge_cases = []
        if examples:
            param_types = infer_param_types(examples)
            edge_cases = generate_edge_cases(examples, constraints, param_types)

        # Combine and deduplicate
        all_cases = examples + edge_cases
        all_cases = deduplicate_cases(all_cases)

        if not all_cases:
            no_examples += 1
            # Still write an empty test file as a placeholder
            all_cases = []

        if edge_cases:
            with_edge += 1

        test_data = {
            "problem_id": int(num),
            "title": title,
            "difficulty": difficulty,
            "link": link,
            "test_cases": all_cases,
        }

        out_path = os.path.join(dir_path, "tests.json")
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(test_data, f, indent=2, ensure_ascii=False, default=str)

        success += 1
        if total % 500 == 0:
            print(f"  Processed {total} problems... ({success} with tests, {with_edge} with edge cases)")

    print(f"\nDone!")
    print(f"  Total non-premium: {total}")
    print(f"  Test files written: {success}")
    print(f"  With parsed examples: {success - no_examples}")
    print(f"  With generated edge cases: {with_edge}")
    print(f"  No examples found: {no_examples}")


if __name__ == "__main__":
    main()
