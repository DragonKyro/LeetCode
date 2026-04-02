"""
Batch Solution Writer

Writes solution files for LeetCode problems from a solutions registry.
The registry maps problem_id -> {python3: code, java: code, cpp: code}.

This script is the driver; actual solutions are loaded from solutions_data/*.json
files which contain the solution code for each problem.

Usage:
    python write_solutions.py                  # write all available solutions
    python write_solutions.py --check          # show stats only
"""

import json
import os
import re
import sys
import io
import argparse
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

PROJECT_ROOT = Path(__file__).resolve().parent
SOLUTIONS_DIR = PROJECT_ROOT / "solutions_data"
PROBLEMS_DIR = PROJECT_ROOT / "problems"


def load_all_solutions():
    """Load all solution data files from solutions_data/."""
    solutions = {}
    if not SOLUTIONS_DIR.exists():
        return solutions
    for f in sorted(SOLUTIONS_DIR.glob("*.json")):
        with open(f, "r", encoding="utf-8") as fh:
            data = json.load(fh)
            for pid_str, sol in data.items():
                solutions[int(pid_str)] = sol
    return solutions


def get_problem_dir(pid):
    """Find the problem directory for a given problem ID."""
    for d in os.listdir(PROBLEMS_DIR):
        if d.startswith(f"{pid}_"):
            return PROBLEMS_DIR / d
    return None


def write_python(problem_dir, code):
    filepath = problem_dir / "solution.py"
    imports = set()
    if "List[" in code:
        imports.add("List")
    if "Optional[" in code:
        imports.add("Optional")
    if "Dict[" in code:
        imports.add("Dict")
    if "Tuple[" in code:
        imports.add("Tuple")

    content = ""
    if imports:
        content += f"from typing import {', '.join(sorted(imports))}\n"
    if "ListNode" in code:
        content += "from shared.python.data_structures import ListNode\n"
    if "TreeNode" in code:
        content += "from shared.python.data_structures import TreeNode\n"
    if imports or "ListNode" in code or "TreeNode" in code:
        content += "\n\n"
    content += code.strip() + "\n"
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)


def write_java(problem_dir, code):
    filepath = problem_dir / "Solution.java"
    imports = ""
    if any(kw in code for kw in ["HashMap", "ArrayList", "List<", "Map<", "Set<",
                                   "Queue<", "Stack<", "Arrays", "Collections",
                                   "PriorityQueue", "LinkedList", "Deque"]):
        imports = "import java.util.*;\n\n"
    content = imports + code.strip() + "\n"
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)


def write_cpp(problem_dir, code):
    filepath = problem_dir / "solution.cpp"
    includes = """#include <vector>
#include <string>
#include <unordered_map>
#include <unordered_set>
#include <algorithm>
#include <climits>
#include <queue>
#include <stack>
#include <cmath>
#include <numeric>
#include <set>
#include <map>
using namespace std;

"""
    content = includes + code.strip() + "\n"
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)


def write_sql(problem_dir, code):
    filepath = problem_dir / "solution.sql"
    content = code.strip() + "\n"
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(content)


def ensure_sql_problem_dir(pid, title):
    """Create problem directory for SQL problems if it doesn't exist."""
    slug = re.sub(r'[^a-z0-9]+', '_', title.lower()).strip('_')
    dir_name = f"{pid}_{slug}"
    dir_path = PROBLEMS_DIR / dir_name
    if not dir_path.exists():
        dir_path.mkdir(parents=True, exist_ok=True)
        # Create a minimal problem.md
        md = f"# {pid}. {title}\n\n**Topics:** Database\n\n---\n\n## Description\n\nSQL Problem\n"
        (dir_path / "problem.md").write_text(md, encoding="utf-8")
    return dir_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    solutions = load_all_solutions()
    print(f"Loaded {len(solutions)} problem solutions")

    if args.check:
        py = sum(1 for s in solutions.values() if "python3" in s)
        java = sum(1 for s in solutions.values() if "java" in s)
        cpp = sum(1 for s in solutions.values() if "cpp" in s)
        sql = sum(1 for s in solutions.values() if "sql" in s)
        print(f"  Python: {py}")
        print(f"  Java:   {java}")
        print(f"  C++:    {cpp}")
        print(f"  SQL:    {sql}")
        return

    written = {"python3": 0, "java": 0, "cpp": 0, "sql": 0}
    missing_dir = 0

    # Load problem titles for SQL dir creation
    from openpyxl import load_workbook
    wb = load_workbook(PROJECT_ROOT / "leetcode_problems.xlsx")
    ws = wb.active
    titles = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        titles[int(row[0])] = row[1]

    for pid, sol in sorted(solutions.items()):
        # For SQL problems, ensure directory exists
        if "sql" in sol and sol["sql"].strip():
            title = titles.get(pid, f"Problem {pid}")
            pdir = ensure_sql_problem_dir(pid, title)
            write_sql(pdir, sol["sql"])
            written["sql"] += 1
            continue

        pdir = get_problem_dir(pid)
        if not pdir or not pdir.exists():
            missing_dir += 1
            continue

        if "python3" in sol and sol["python3"].strip():
            write_python(pdir, sol["python3"])
            written["python3"] += 1
        if "java" in sol and sol["java"].strip():
            write_java(pdir, sol["java"])
            written["java"] += 1
        if "cpp" in sol and sol["cpp"].strip():
            write_cpp(pdir, sol["cpp"])
            written["cpp"] += 1

    print(f"\nWritten:")
    print(f"  Python: {written['python3']}")
    print(f"  Java:   {written['java']}")
    print(f"  C++:    {written['cpp']}")
    print(f"  SQL:    {written['sql']}")
    if missing_dir:
        print(f"  Missing dirs: {missing_dir}")


if __name__ == "__main__":
    main()
