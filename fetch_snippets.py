"""
Fetch LeetCode code snippets (function signatures) for Python, Java, C++.
Saves to snippets.json keyed by problem number.
"""

import requests
import time
import json
import sys
import io
import os

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

GRAPHQL_URL = "https://leetcode.com/graphql"
HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://leetcode.com/problemset/",
}

CHECKPOINT = "snippets_checkpoint.json"
OUTPUT = "snippets.json"


def load_checkpoint():
    try:
        with open(CHECKPOINT, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"snippets": {}, "last_index": -1}


def save_checkpoint(data, idx):
    with open(CHECKPOINT, "w", encoding="utf-8") as f:
        json.dump({"snippets": data, "last_index": idx}, f, ensure_ascii=False)


def fetch_snippets(title_slug):
    query = {
        "query": """
        query questionData($titleSlug: String!) {
            question(titleSlug: $titleSlug) {
                questionFrontendId
                titleSlug
                codeSnippets {
                    lang
                    langSlug
                    code
                }
            }
        }
        """,
        "variables": {"titleSlug": title_slug},
    }
    resp = requests.post(GRAPHQL_URL, json=query, headers=HEADERS)
    resp.raise_for_status()
    return resp.json()["data"]["question"]


def main():
    from openpyxl import load_workbook

    print("Fetching code snippets from LeetCode...")

    wb = load_workbook("leetcode_problems.xlsx")
    ws = wb.active

    # Build list of non-premium problems
    problems = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, title, diff, desc, topics, hints, link, premium = row
        if premium == "Yes":
            continue
        # Check if problem directory exists
        import re
        slug_dir = re.sub(r'[^a-z0-9]+', '_', title.lower()).strip('_')
        dir_name = f"{int(num)}_{slug_dir}"
        if not os.path.exists(os.path.join("problems", dir_name)):
            continue
        # Extract titleSlug from link
        title_slug = link.rstrip("/").split("/")[-1]
        problems.append({"num": int(num), "title": title, "slug": title_slug, "dir": dir_name})

    print(f"  {len(problems)} problems to fetch")

    # Load checkpoint
    cp = load_checkpoint()
    snippets = cp["snippets"]
    start = cp["last_index"] + 1

    if start > 0:
        print(f"  Resuming from index {start}")

    for i in range(start, len(problems)):
        prob = problems[i]
        try:
            data = fetch_snippets(prob["slug"])
            if data and data.get("codeSnippets"):
                entry = {}
                for snip in data["codeSnippets"]:
                    if snip["langSlug"] in ("python3", "java", "cpp"):
                        entry[snip["langSlug"]] = snip["code"]
                snippets[str(prob["num"])] = {
                    "title": prob["title"],
                    "dir": prob["dir"],
                    "slug": prob["slug"],
                    "snippets": entry,
                }
            print(f"  [{i+1}/{len(problems)}] #{prob['num']} {prob['title']}")
        except Exception as e:
            print(f"  [{i+1}/{len(problems)}] #{prob['num']} ERROR: {e}")

        if (i + 1) % 50 == 0:
            save_checkpoint(snippets, i)

        time.sleep(0.5)

    # Save final
    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(snippets, f, indent=2, ensure_ascii=False)

    if os.path.exists(CHECKPOINT):
        os.remove(CHECKPOINT)

    print(f"\nDone! Saved {len(snippets)} snippets to {OUTPUT}")


if __name__ == "__main__":
    main()
