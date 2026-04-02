"""
LeetCode Problem Scraper
Scrapes all LeetCode problems using the public GraphQL API and saves to Excel.

For each problem: number, title, description (with examples/constraints),
difficulty, topics, hints, and link. Premium-gated problems are noted.
"""

import requests
import time
import html
import re
import json
import sys
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Fix Windows console encoding for Unicode characters
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

GRAPHQL_URL = "https://leetcode.com/graphql"

HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://leetcode.com/problemset/",
}


def fetch_problem_list():
    """Fetch the full list of problems using the problemsetQuestionList query."""
    all_problems = []
    skip = 0
    limit = 100  # max per request

    while True:
        query = {
            "query": """
            query problemsetQuestionList($categorySlug: String, $limit: Int, $skip: Int, $filters: QuestionListFilterInput) {
                problemsetQuestionList: questionList(
                    categorySlug: $categorySlug
                    limit: $limit
                    skip: $skip
                    filters: $filters
                ) {
                    total: totalNum
                    questions: data {
                        frontendQuestionId: questionFrontendId
                        title
                        titleSlug
                        difficulty
                        paidOnly: isPaidOnly
                        topicTags {
                            name
                        }
                    }
                }
            }
            """,
            "variables": {
                "categorySlug": "",
                "skip": skip,
                "limit": limit,
                "filters": {},
            },
        }

        resp = requests.post(GRAPHQL_URL, json=query, headers=HEADERS)
        resp.raise_for_status()
        data = resp.json()

        question_list = data["data"]["problemsetQuestionList"]
        total = question_list["total"]
        questions = question_list["questions"]

        if not questions:
            break

        all_problems.extend(questions)
        skip += limit

        print(f"  Fetched {len(all_problems)}/{total} problems...")

        if len(all_problems) >= total:
            break

        time.sleep(0.5)  # be polite

    return all_problems


def strip_html(text):
    """Convert HTML content to plain text, preserving structure."""
    if not text:
        return ""
    # Replace common HTML tags with text equivalents
    text = re.sub(r"<br\s*/?>", "\n", text)
    text = re.sub(r"<p>", "\n", text)
    text = re.sub(r"</p>", "", text)
    text = re.sub(r"<li>", "\n• ", text)
    text = re.sub(r"</li>", "", text)
    text = re.sub(r"<ul>|</ul>|<ol>|</ol>", "\n", text)
    text = re.sub(r"<strong>|</strong>|<b>|</b>", "", text)
    text = re.sub(r"<em>|</em>|<i>|</i>", "", text)
    text = re.sub(r"<code>|</code>", "", text)
    text = re.sub(r"<pre>|</pre>", "\n", text)
    text = re.sub(r"<sup>", "^", text)
    text = re.sub(r"</sup>", "", text)
    text = re.sub(r"<sub>", "_", text)
    text = re.sub(r"</sub>", "", text)
    # Remove image tags but keep alt text
    text = re.sub(r'<img[^>]*alt="([^"]*)"[^>]*>', r"\1", text)
    text = re.sub(r"<img[^>]*>", "", text)
    # Remove any remaining tags
    text = re.sub(r"<[^>]+>", "", text)
    text = html.unescape(text)
    # Clean up excessive whitespace
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def fetch_problem_detail(title_slug):
    """Fetch detailed info for a single problem."""
    query = {
        "query": """
        query questionData($titleSlug: String!) {
            question(titleSlug: $titleSlug) {
                questionFrontendId
                title
                content
                difficulty
                hints
                topicTags {
                    name
                }
                isPaidOnly
            }
        }
        """,
        "variables": {"titleSlug": title_slug},
    }

    resp = requests.post(GRAPHQL_URL, json=query, headers=HEADERS)
    resp.raise_for_status()
    data = resp.json()
    return data["data"]["question"]


def save_to_excel(problems, filename="leetcode_problems.xlsx"):
    """Save all problem data to an Excel spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LeetCode Problems"

    # Define headers
    headers = [
        "Problem #",
        "Title",
        "Difficulty",
        "Description",
        "Topics",
        "Hints",
        "Link",
        "Premium Only",
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    easy_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    hard_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    premium_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Sort problems by number
    problems.sort(key=lambda x: int(x.get("number", 0)))

    # Write data
    for row_idx, prob in enumerate(problems, 2):
        num = prob.get("number", "")
        title = prob.get("title", "")
        difficulty = prob.get("difficulty", "")
        description = prob.get("description", "")
        topics = prob.get("topics", "")
        hints = prob.get("hints", "")
        link = prob.get("link", "")
        premium = prob.get("premium", False)

        ws.cell(row=row_idx, column=1, value=int(num) if num else "").border = thin_border
        ws.cell(row=row_idx, column=2, value=title).border = thin_border
        diff_cell = ws.cell(row=row_idx, column=3, value=difficulty)
        diff_cell.border = thin_border
        ws.cell(row=row_idx, column=4, value=description).border = thin_border
        ws.cell(row=row_idx, column=5, value=topics).border = thin_border
        ws.cell(row=row_idx, column=6, value=hints).border = thin_border
        ws.cell(row=row_idx, column=7, value=link).border = thin_border
        premium_cell = ws.cell(row=row_idx, column=8, value="Yes" if premium else "No")
        premium_cell.border = thin_border

        # Color-code difficulty
        if difficulty == "Easy":
            diff_cell.fill = easy_fill
        elif difficulty == "Medium":
            diff_cell.fill = medium_fill
        elif difficulty == "Hard":
            diff_cell.fill = hard_fill

        if premium:
            premium_cell.fill = premium_fill

        # Wrap text for long fields
        for col in [4, 5, 6]:
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    # Set column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 80
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["G"].width = 50
    ws.column_dimensions["H"].width = 14

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    print(f"\nSaved {len(problems)} problems to {filename}")


CHECKPOINT_FILE = "scrape_checkpoint.json"


def load_checkpoint():
    """Load previously scraped data from checkpoint file."""
    try:
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"problems": [], "last_index": -1}


def save_checkpoint(problems, last_index):
    """Save progress to checkpoint file."""
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump({"problems": problems, "last_index": last_index}, f, ensure_ascii=False)


def main():
    print("=" * 60)
    print("LeetCode Problem Scraper")
    print("=" * 60)

    # Step 1: Fetch problem list
    print("\n[1/3] Fetching problem list...")
    raw_problems = fetch_problem_list()
    print(f"  Found {len(raw_problems)} problems total.")

    # Check for checkpoint to resume
    checkpoint = load_checkpoint()
    problems = checkpoint["problems"]
    start_index = checkpoint["last_index"] + 1

    if start_index > 0:
        print(f"\n  Resuming from index {start_index} (already have {len(problems)} problems)")

    # Step 2: Fetch details for each non-premium problem
    print("\n[2/3] Fetching problem details...")
    premium_count = sum(1 for p in problems if p.get("premium"))
    error_count = 0

    for i in range(start_index, len(raw_problems)):
        prob = raw_problems[i]
        num = prob["frontendQuestionId"]
        title = prob["title"]
        slug = prob["titleSlug"]
        difficulty = prob["difficulty"]
        is_premium = prob["paidOnly"]
        topics = ", ".join(tag["name"] for tag in prob.get("topicTags", []))
        link = f"https://leetcode.com/problems/{slug}/"

        if is_premium:
            # Record premium problems with limited info
            problems.append({
                "number": num,
                "title": title,
                "difficulty": difficulty,
                "description": "[Premium - content not available]",
                "topics": topics,
                "hints": "",
                "link": link,
                "premium": True,
            })
            premium_count += 1
            print(f"  [{i+1}/{len(raw_problems)}] #{num} {title} (Premium - skipping detail)")
        else:
            # Fetch full details
            try:
                detail = fetch_problem_detail(slug)
                description = strip_html(detail.get("content", "")) if detail.get("content") else ""
                hints_list = detail.get("hints", []) or []
                hints_text = "\n".join(
                    f"Hint {j+1}: {strip_html(h)}" for j, h in enumerate(hints_list)
                )
                # Use topics from detail if available (more complete)
                detail_topics = ", ".join(
                    tag["name"] for tag in detail.get("topicTags", [])
                )

                problems.append({
                    "number": num,
                    "title": title,
                    "difficulty": difficulty,
                    "description": description,
                    "topics": detail_topics or topics,
                    "hints": hints_text,
                    "link": link,
                    "premium": False,
                })
                print(f"  [{i+1}/{len(raw_problems)}] #{num} {title}")
            except Exception as e:
                # If detail fetch fails, still record basic info
                problems.append({
                    "number": num,
                    "title": title,
                    "difficulty": difficulty,
                    "description": f"[Error fetching details: {e}]",
                    "topics": topics,
                    "hints": "",
                    "link": link,
                    "premium": False,
                })
                error_count += 1
                print(f"  [{i+1}/{len(raw_problems)}] #{num} {title} (ERROR: {e})")

            # Rate limiting - be polite to the API
            time.sleep(0.5)

        # Save checkpoint every 50 problems
        if (i + 1) % 50 == 0:
            save_checkpoint(problems, i)

    # Step 3: Save to Excel
    print(f"\n[3/3] Saving to Excel...")
    print(f"  Total: {len(problems)} | Free: {len(problems) - premium_count} | Premium: {premium_count} | Errors: {error_count}")
    save_to_excel(problems, "leetcode_problems.xlsx")

    # Clean up checkpoint file
    import os
    if os.path.exists(CHECKPOINT_FILE):
        os.remove(CHECKPOINT_FILE)

    print("\nDone!")


if __name__ == "__main__":
    main()
